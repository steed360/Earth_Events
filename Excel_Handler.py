from openpyxl import Workbook
import openpyxl
import os

excel_file_path =os.path.join (os.getcwd(),'EONET_Events.xlsx' )

def createWorkbook ():


    if (os.path.exists (excel_file_path) ):
        os.remove(  excel_file_path)

    wb = Workbook()
    wb.create_sheet('EONET_Events')
    sheet = wb.get_sheet_by_name('EONET_Events') 

    sheet['A2'] = 'EONET Events in the last 30 days'

    sheet['A4'] = 'Event ID'
    sheet['B4'] = 'Description'
    sheet['C4'] = '-'
    sheet['D4'] = 'Link'
    sheet['E4'] = 'Categories'
    sheet['F4'] = 'Country (where obtainable)'

    wb.save(  excel_file_path)
    wb.close()


def writeEvents (eventsList):
    # eventsList = [('a','b'), ('b', 'c')]
    wb = openpyxl.load_workbook( excel_file_path)
    ws = wb.get_sheet_by_name ( 'EONET_Events')
    for row in eventsList:
       ws.append(row)
    wb.save (excel_file_path)
    wb.close()

def getExcelFilePath ():
    return excel_file_path

if __name__ == '__main__' :
    createWorkbook ()
    import JSON_handler
    d = JSON_handler.getEventsDict ()
    writeEvents (d)


